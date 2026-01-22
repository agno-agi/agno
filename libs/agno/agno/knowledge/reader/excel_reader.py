import asyncio
import io
from pathlib import Path
from typing import IO, Any, Iterable, List, Optional, Sequence, Tuple, Union

from agno.knowledge.chunking.row import RowChunking
from agno.knowledge.chunking.strategy import ChunkingStrategy, ChunkingStrategyType
from agno.knowledge.document.base import Document
from agno.knowledge.reader.base import Reader
from agno.knowledge.reader.spreadsheet_utils import (
    convert_xls_cell_value,
    excel_rows_to_documents,
    get_workbook_name,
    infer_file_extension,
)
from agno.knowledge.types import ContentType
from agno.utils.log import log_debug, log_error


class ExcelReader(Reader):
    """Reader for Excel files (.xlsx and .xls).

    Converts Excel workbooks to documents, with one document per sheet by default.
    Supports filtering to specific sheets and optional chunking.

    Args:
        sheets: List of sheet names or 0-based indices to include. None = all sheets.
        include_empty_sheets: Whether to include sheets with no data. Default False.
        skip_hidden_sheets: Whether to skip hidden sheets (xlsx only). Default True.
        chunking_strategy: Strategy for chunking documents. Default is RowChunking.
        **kwargs: Additional arguments passed to base Reader.
    """

    def __init__(
        self,
        sheets: Optional[List[Union[str, int]]] = None,
        include_empty_sheets: bool = False,
        skip_hidden_sheets: bool = True,
        chunking_strategy: Optional[ChunkingStrategy] = RowChunking(),
        **kwargs,
    ):
        super().__init__(chunking_strategy=chunking_strategy, **kwargs)
        self.sheets = sheets
        self.include_empty_sheets = include_empty_sheets
        self.skip_hidden_sheets = skip_hidden_sheets

    @classmethod
    def get_supported_chunking_strategies(cls) -> List[ChunkingStrategyType]:
        """Get the list of supported chunking strategies for Excel readers."""
        return [
            ChunkingStrategyType.ROW_CHUNKER,
            ChunkingStrategyType.CODE_CHUNKER,
            ChunkingStrategyType.FIXED_SIZE_CHUNKER,
            ChunkingStrategyType.AGENTIC_CHUNKER,
            ChunkingStrategyType.DOCUMENT_CHUNKER,
            ChunkingStrategyType.RECURSIVE_CHUNKER,
        ]

    @classmethod
    def get_supported_content_types(cls) -> List[ContentType]:
        """Get the list of supported content types."""
        return [ContentType.XLSX, ContentType.XLS]

    def _should_include_sheet(
        self,
        sheet_name: str,
        sheet_index: int,
        is_hidden: bool = False,
    ) -> bool:
        """Check if sheet passes the configured filters."""
        if is_hidden and self.skip_hidden_sheets:
            return False

        if self.sheets is None:
            return True

        for sheet_filter in self.sheets:
            if isinstance(sheet_filter, int):
                if sheet_index == sheet_filter:
                    return True
            elif isinstance(sheet_filter, str):
                if sheet_name == sheet_filter:
                    return True

        return False

    def _read_xlsx(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        """Read .xlsx file using openpyxl."""
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                "`openpyxl` not installed. Please install it via `pip install agno[excel]` or `pip install openpyxl`."
            ) from e

        if isinstance(file, Path):
            workbook = openpyxl.load_workbook(filename=str(file), read_only=True, data_only=True)
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = openpyxl.load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)

        try:
            sheets: List[Tuple[str, Iterable[Sequence[Any]]]] = []
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                # Check visibility (xlsx supports hidden sheets)
                is_hidden = worksheet.sheet_state != "visible"

                if not self._should_include_sheet(worksheet.title, sheet_index, is_hidden):
                    log_debug(f"Skipping sheet '{worksheet.title}' (filtered out)")
                    continue

                sheets.append((worksheet.title, worksheet.iter_rows(values_only=True)))

            return excel_rows_to_documents(workbook_name=workbook_name, sheets=sheets)
        finally:
            workbook.close()

    def _read_xls(self, file: Union[Path, IO[Any]], *, workbook_name: str) -> List[Document]:
        """Read .xls file using xlrd."""
        try:
            import xlrd
        except ImportError as e:
            raise ImportError(
                "`xlrd` not installed. Please install it via `pip install agno[excel]` or `pip install xlrd`."
            ) from e

        if isinstance(file, Path):
            workbook = xlrd.open_workbook(filename=str(file))
        else:
            file.seek(0)
            raw = file.read()
            if isinstance(raw, str):
                raw = raw.encode("utf-8", errors="replace")
            workbook = xlrd.open_workbook(file_contents=raw)

        sheets: List[Tuple[str, Iterable[Sequence[Any]]]] = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)

            # Note: xlrd doesn't expose hidden sheet info easily, so we skip that check
            if not self._should_include_sheet(sheet.name, sheet_index, is_hidden=False):
                log_debug(f"Skipping sheet '{sheet.name}' (filtered out)")
                continue

            def _iter_sheet_rows(_sheet: Any = sheet, _datemode: int = workbook.datemode) -> Iterable[Sequence[Any]]:
                for row_index in range(_sheet.nrows):
                    yield [
                        convert_xls_cell_value(
                            _sheet.cell_value(row_index, col_index),
                            _sheet.cell_type(row_index, col_index),
                            _datemode,
                        )
                        for col_index in range(_sheet.ncols)
                    ]

            sheets.append((sheet.name, _iter_sheet_rows()))

        return excel_rows_to_documents(workbook_name=workbook_name, sheets=sheets)

    def read(
        self,
        file: Union[Path, IO[Any]],
        name: Optional[str] = None,
    ) -> List[Document]:
        """Read an Excel file and return documents (one per sheet)."""
        try:
            file_extension = infer_file_extension(file, name)
            workbook_name = get_workbook_name(file, name)

            if isinstance(file, Path) and not file.exists():
                raise FileNotFoundError(f"Could not find file: {file}")

            file_desc = str(file) if isinstance(file, Path) else getattr(file, "name", "BytesIO")
            log_debug(f"Reading Excel file: {file_desc}")

            if file_extension == ContentType.XLSX or file_extension == ".xlsx":
                documents = self._read_xlsx(file, workbook_name=workbook_name)
            elif file_extension == ContentType.XLS or file_extension == ".xls":
                documents = self._read_xls(file, workbook_name=workbook_name)
            else:
                log_error(f"Unsupported file extension: {file_extension}. Expected .xlsx or .xls")
                return []

            if self.chunk:
                chunked_documents = []
                for document in documents:
                    chunked_documents.extend(self.chunk_document(document))
                return chunked_documents

            return documents

        except FileNotFoundError:
            raise
        except ImportError:
            raise
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []

    async def async_read(
        self,
        file: Union[Path, IO[Any]],
        name: Optional[str] = None,
    ) -> List[Document]:
        """Async version of read()."""
        try:
            file_extension = infer_file_extension(file, name)
            workbook_name = get_workbook_name(file, name)

            if isinstance(file, Path) and not file.exists():
                raise FileNotFoundError(f"Could not find file: {file}")

            file_desc = str(file) if isinstance(file, Path) else getattr(file, "name", "BytesIO")
            log_debug(f"Reading Excel file async: {file_desc}")

            if file_extension == ContentType.XLSX or file_extension == ".xlsx":
                documents = await asyncio.to_thread(self._read_xlsx, file, workbook_name=workbook_name)
            elif file_extension == ContentType.XLS or file_extension == ".xls":
                documents = await asyncio.to_thread(self._read_xls, file, workbook_name=workbook_name)
            else:
                log_error(f"Unsupported file extension: {file_extension}. Expected .xlsx or .xls")
                return []

            if self.chunk:
                documents = await self.chunk_documents_async(documents)

            return documents

        except FileNotFoundError:
            raise
        except ImportError:
            raise
        except Exception as e:
            file_desc = getattr(file, "name", str(file)) if isinstance(file, IO) else file
            log_error(f"Error reading {file_desc}: {e}")
            return []
