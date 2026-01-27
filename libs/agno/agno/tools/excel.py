"""Excel Tools for creating and manipulating Excel files.

This toolkit provides AI agents with the ability to create, read, and modify
Excel workbooks (.xlsx format) using the openpyxl library.

Install dependencies: pip install openpyxl
"""

import json
from pathlib import Path
from typing import Any, List, Optional, Tuple

from agno.tools import Toolkit
from agno.utils.log import log_debug, log_error, log_info


class ExcelTools(Toolkit):
    """A toolkit for creating and manipulating Excel files.

    This toolkit enables AI agents to:
    - Create new Excel workbooks
    - Write data to cells and ranges
    - Read data from sheets
    - Add and manage sheets
    - Add formulas to cells
    - Apply cell formatting

    Example:
        ```python
        from agno.agent import Agent
        from agno.models.openai import OpenAIChat
        from agno.tools.excel import ExcelTools

        agent = Agent(
            model=OpenAIChat(id="gpt-4o"),
            tools=[ExcelTools()],
        )
        agent.print_response("Create an Excel file with sales data")
        ```
    """

    def __init__(
        self,
        base_dir: Optional[Path] = None,
        enable_create_workbook: bool = True,
        enable_write_data: bool = True,
        enable_read_data: bool = True,
        enable_add_sheet: bool = True,
        enable_list_sheets: bool = True,
        enable_get_file_path: bool = True,
        enable_add_formula: bool = False,
        enable_format_cells: bool = False,
        all: bool = False,
        **kwargs: Any,
    ):
        """Initialize the Excel toolkit.

        Args:
            base_dir: Base directory for Excel file operations. If not provided,
                uses the current working directory.
            enable_create_workbook: Enable the create_workbook tool. Default: True.
            enable_write_data: Enable the write_data tool. Default: True.
            enable_read_data: Enable the read_data tool. Default: True.
            enable_add_sheet: Enable the add_sheet tool. Default: True.
            enable_list_sheets: Enable the list_sheets tool. Default: True.
            enable_get_file_path: Enable the get_file_path tool. Default: True.
            enable_add_formula: Enable the add_formula tool. Default: False.
            enable_format_cells: Enable the format_cells tool. Default: False.
            all: Enable all tools. Default: False.
            **kwargs: Additional arguments passed to the Toolkit base class.
        """
        self.base_dir: Path = (base_dir or Path.cwd()).resolve()

        tools: List[Any] = []
        if all or enable_create_workbook:
            tools.append(self.create_workbook)
        if all or enable_write_data:
            tools.append(self.write_data)
        if all or enable_read_data:
            tools.append(self.read_data)
        if all or enable_add_sheet:
            tools.append(self.add_sheet)
        if all or enable_list_sheets:
            tools.append(self.list_sheets)
        if all or enable_get_file_path:
            tools.append(self.get_file_path)
        if all or enable_add_formula:
            tools.append(self.add_formula)
        if all or enable_format_cells:
            tools.append(self.format_cells)

        super().__init__(name="excel_tools", tools=tools, **kwargs)

    def _check_path(self, file_name: str) -> Tuple[bool, Path]:
        """Check if the file path is within the base directory.

        Args:
            file_name: The file name or relative path to check.

        Returns:
            Tuple of (is_safe, resolved_path). If not safe, returns base_dir as path.
        """
        # Delegate to parent Toolkit class method
        return super()._check_path(file_name, self.base_dir)

    def _ensure_xlsx_extension(self, file_name: str) -> str:
        """Ensure the file has .xlsx extension."""
        if not file_name.endswith(".xlsx"):
            return f"{file_name}.xlsx"
        return file_name

    def _get_openpyxl(self):
        """Import and return openpyxl, raising ImportError if not installed."""
        try:
            import openpyxl

            return openpyxl
        except ImportError as e:
            raise ImportError(
                "`openpyxl` not installed. Please install it via `pip install openpyxl`."
            ) from e

    def create_workbook(
        self,
        file_name: str,
        sheet_name: Optional[str] = None,
        overwrite: bool = False,
    ) -> str:
        """Create a new Excel workbook.

        Args:
            file_name: Name of the Excel file to create (will add .xlsx if missing).
            sheet_name: Optional name for the first sheet. Defaults to 'Sheet'.
            overwrite: If True, overwrite existing file. Default: False.

        Returns:
            JSON string with the result including file path and sheet name.
        """
        try:
            openpyxl = self._get_openpyxl()

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to create file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if file_path.exists() and not overwrite:
                return json.dumps(
                    {
                        "error": f"File '{file_name}' already exists. Set overwrite=True to replace."
                    }
                )

            log_info(f"Creating Excel workbook: {file_path}")

            # Create new workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            if sheet_name:
                sheet.title = sheet_name

            # Ensure parent directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            workbook.save(str(file_path))
            workbook.close()

            log_debug(f"Created workbook: {file_path}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheet": sheet.title,
                    "message": f"Created Excel workbook '{file_name}' with sheet '{sheet.title}'",
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error creating workbook: {e}")
            return json.dumps({"error": f"Error creating workbook: {e}"})

    def write_data(
        self,
        file_name: str,
        data: List[List[Any]],
        sheet_name: Optional[str] = None,
        start_cell: str = "A1",
    ) -> str:
        """Write data to an Excel sheet.

        Args:
            file_name: Name of the Excel file.
            data: 2D list of data to write (rows and columns).
            sheet_name: Name of the sheet to write to. Uses active sheet if not specified.
            start_cell: Starting cell for data (e.g., "A1", "B2"). Default: "A1".

        Returns:
            JSON string with the result including rows and columns written.
        """
        try:
            openpyxl = self._get_openpyxl()
            from openpyxl.utils.cell import (
                column_index_from_string,
                coordinate_from_string,
            )

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps(
                    {
                        "error": f"File '{file_name}' not found. Create it first with create_workbook."
                    }
                )

            log_info(f"Writing data to: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path))

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return json.dumps(
                        {
                            "error": f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
                        }
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Parse start cell
            col_letter, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)

            # Write data
            rows_written = 0
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    sheet.cell(
                        row=start_row + row_idx, column=start_col + col_idx, value=value
                    )
                rows_written += 1

            workbook.save(str(file_path))
            workbook.close()

            log_debug(f"Wrote {rows_written} rows to {sheet.title}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheet": sheet.title,
                    "rows_written": rows_written,
                    "columns_written": len(data[0]) if data else 0,
                    "start_cell": start_cell,
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error writing data: {e}")
            return json.dumps({"error": f"Error writing data: {e}"})

    def read_data(
        self,
        file_name: str,
        sheet_name: Optional[str] = None,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
    ) -> str:
        """Read data from an Excel sheet.

        Args:
            file_name: Name of the Excel file.
            sheet_name: Name of the sheet to read from. Uses active sheet if not specified.
            start_cell: Starting cell to read from (e.g., "A1"). Default: "A1".
            end_cell: Optional ending cell. If not specified, reads all data from start.

        Returns:
            JSON string with the data as a 2D list.
        """
        try:
            openpyxl = self._get_openpyxl()

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps({"error": f"File '{file_name}' not found."})

            log_info(f"Reading data from: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path), data_only=True)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    workbook.close()
                    return json.dumps(
                        {
                            "error": f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
                        }
                    )
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Define range
            if end_cell:
                cell_range = f"{start_cell}:{end_cell}"
            else:
                # Get the max row and column from the sheet
                max_row = sheet.max_row or 1
                max_col = sheet.max_column or 1
                from openpyxl.utils import get_column_letter

                end_col_letter = get_column_letter(max_col)
                cell_range = f"{start_cell}:{end_col_letter}{max_row}"

            # Read data
            data = []
            for row in sheet[cell_range]:
                row_data = []
                for cell in row:
                    value = cell.value
                    # Convert datetime objects to string for JSON serialization
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    row_data.append(value)
                data.append(row_data)

            workbook.close()

            log_debug(f"Read {len(data)} rows from {sheet.title}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheet": sheet.title,
                    "data": data,
                    "rows": len(data),
                    "columns": len(data[0]) if data else 0,
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error reading data: {e}")
            return json.dumps({"error": f"Error reading data: {e}"})

    def add_sheet(
        self,
        file_name: str,
        sheet_name: str,
        position: Optional[int] = None,
    ) -> str:
        """Add a new sheet to an existing workbook.

        Args:
            file_name: Name of the Excel file.
            sheet_name: Name for the new sheet.
            position: Optional position for the sheet (0-indexed). Adds at end if not specified.

        Returns:
            JSON string with the result including all sheet names.
        """
        try:
            openpyxl = self._get_openpyxl()

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps(
                    {
                        "error": f"File '{file_name}' not found. Create it first with create_workbook."
                    }
                )

            log_info(f"Adding sheet '{sheet_name}' to: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path))

            if sheet_name in workbook.sheetnames:
                return json.dumps({"error": f"Sheet '{sheet_name}' already exists."})

            if position is not None:
                workbook.create_sheet(title=sheet_name, index=position)
            else:
                workbook.create_sheet(title=sheet_name)

            workbook.save(str(file_path))
            sheets = workbook.sheetnames
            workbook.close()

            log_debug(f"Added sheet '{sheet_name}' to {file_name}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "new_sheet": sheet_name,
                    "all_sheets": sheets,
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error adding sheet: {e}")
            return json.dumps({"error": f"Error adding sheet: {e}"})

    def list_sheets(self, file_name: str) -> str:
        """List all sheets in a workbook.

        Args:
            file_name: Name of the Excel file.

        Returns:
            JSON string with the list of sheet names.
        """
        try:
            openpyxl = self._get_openpyxl()

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps({"error": f"File '{file_name}' not found."})

            log_debug(f"Listing sheets in: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path), read_only=True)
            sheets = workbook.sheetnames
            workbook.close()

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheets": sheets,
                    "count": len(sheets),
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error listing sheets: {e}")
            return json.dumps({"error": f"Error listing sheets: {e}"})

    def get_file_path(self, file_name: str) -> str:
        """Get the full absolute path to an Excel file.

        Use this to retrieve the file location for downloading or accessing
        the Excel file outside of the agent.

        Args:
            file_name: Name of the Excel file.

        Returns:
            JSON string with the full file path and file info.
        """
        try:
            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps({"error": f"File '{file_name}' not found."})

            # Get file size
            file_size = file_path.stat().st_size

            log_debug(f"Retrieved path for: {file_path}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "path": str(file_path),
                    "size_bytes": file_size,
                    "exists": True,
                }
            )

        except Exception as e:
            log_error(f"Error getting file path: {e}")
            return json.dumps({"error": f"Error getting file path: {e}"})

    def add_formula(
        self,
        file_name: str,
        cell: str,
        formula: str,
        sheet_name: Optional[str] = None,
    ) -> str:
        """Add a formula to a cell.

        Args:
            file_name: Name of the Excel file.
            cell: Cell reference (e.g., "C1", "D5").
            formula: Excel formula (e.g., "=SUM(A1:B1)", "=A1*B1").
            sheet_name: Name of the sheet. Uses active sheet if not specified.

        Returns:
            JSON string with the result.
        """
        try:
            openpyxl = self._get_openpyxl()

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps({"error": f"File '{file_name}' not found."})

            # Ensure formula starts with =
            if not formula.startswith("="):
                formula = f"={formula}"

            log_info(f"Adding formula to {cell} in: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path))

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return json.dumps({"error": f"Sheet '{sheet_name}' not found."})
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            sheet[cell] = formula

            workbook.save(str(file_path))
            workbook.close()

            log_debug(f"Added formula '{formula}' to {cell}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheet": sheet.title,
                    "cell": cell,
                    "formula": formula,
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error adding formula: {e}")
            return json.dumps({"error": f"Error adding formula: {e}"})

    def format_cells(
        self,
        file_name: str,
        cell_range: str,
        sheet_name: Optional[str] = None,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        font_size: Optional[int] = None,
        font_color: Optional[str] = None,
        bg_color: Optional[str] = None,
    ) -> str:
        """Apply formatting to cells.

        Args:
            file_name: Name of the Excel file.
            cell_range: Cell or range to format (e.g., "A1", "A1:C5").
            sheet_name: Name of the sheet. Uses active sheet if not specified.
            bold: Set bold formatting.
            italic: Set italic formatting.
            font_size: Font size in points.
            font_color: Font color as hex (e.g., "FF0000" for red).
            bg_color: Background color as hex (e.g., "FFFF00" for yellow).

        Returns:
            JSON string with the result.
        """
        try:
            openpyxl = self._get_openpyxl()
            from openpyxl.styles import Font, PatternFill

            file_name = self._ensure_xlsx_extension(file_name)
            safe, file_path = self._check_path(file_name)

            if not safe:
                log_error(
                    f"Attempted to access file outside base directory: {file_name}"
                )
                return json.dumps({"error": "Invalid file path"})

            if not file_path.exists():
                return json.dumps({"error": f"File '{file_name}' not found."})

            log_info(f"Formatting cells {cell_range} in: {file_path}")

            workbook = openpyxl.load_workbook(str(file_path))

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return json.dumps({"error": f"Sheet '{sheet_name}' not found."})
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            # Build font style
            font_kwargs: dict[str, Any] = {}
            if bold is not None:
                font_kwargs["bold"] = bold
            if italic is not None:
                font_kwargs["italic"] = italic
            if font_size is not None:
                font_kwargs["size"] = font_size
            if font_color is not None:
                font_kwargs["color"] = font_color

            font = Font(**font_kwargs) if font_kwargs else None

            # Build fill style
            fill = None
            if bg_color:
                fill = PatternFill(
                    start_color=bg_color, end_color=bg_color, fill_type="solid"
                )

            # Apply formatting to range
            cells_formatted = 0
            cell_selection = sheet[cell_range]

            # Handle single cell case (returns Cell directly, not tuple)
            from openpyxl.cell import Cell

            if isinstance(cell_selection, Cell):
                if font:
                    cell_selection.font = font
                if fill:
                    cell_selection.fill = fill
                cells_formatted = 1
            else:
                for row in cell_selection:
                    if not hasattr(row, "__iter__"):
                        row = [row]
                    for cell in row:
                        if font:
                            cell.font = font
                        if fill:
                            cell.fill = fill
                        cells_formatted += 1

            workbook.save(str(file_path))
            workbook.close()

            log_debug(f"Formatted {cells_formatted} cells in {cell_range}")

            return json.dumps(
                {
                    "success": True,
                    "file": file_name,
                    "sheet": sheet.title,
                    "range": cell_range,
                    "cells_formatted": cells_formatted,
                }
            )

        except ImportError as e:
            return json.dumps({"error": str(e)})
        except Exception as e:
            log_error(f"Error formatting cells: {e}")
            return json.dumps({"error": f"Error formatting cells: {e}"})
